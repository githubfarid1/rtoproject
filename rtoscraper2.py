import os
import requests
import json
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook, load_workbook
import argparse
import setting as s
from requests import Session
from requests.exceptions import ProxyError
from itertools import groupby
from prettytable import PrettyTable

'''
json.loads : untuk mengubah STRING JSON menjadi OBJECT DICTIONARY

json.dumps: mengubah OBJECT DICTIONARY menjadi STRING JSON
'''
           
cookies = {
    'downloadStarted': 'false',
    'ASP.NET_SessionId': 'x1qkewrzf3bkng5fe5guccs0',
    '__utmc': '185625580',
    '__utmz': '185625580.1709167767.3.3.utmcsr=upwork.com|utmccn=(referral)|utmcmd=referral|utmcct=/',
    'TgaSearchOption': 'Rto',
    'Organisation_AjaxDetailsLoadScope_ScopeQualification': '20',
    'ifShowHistory': 'false',
    '__utma': '185625580.1342116690.1708677775.1709175837.1709180464.5',
    '__utmt': '1',
    '.ASPXANONYMOUS': 'TZ4IATFyAr8sYnN-VyfM1-Nezh4esW6nhqc-NSyNFUuXMeeW9joXb16yxqMxN0OH1brP5IRl7edCR79tQUiYB0nu4yuqYXCcDYcZTkBqtA5dMqO_HnmfiyoluLoeDetDw5TOROO8D3NIXy54RoZllw2',
    '__utmb': '185625580.5.10.1709180464',
    'Search_Index_gridRtoSearchResults': '100',
}

headers = {
    'authority': 'training.gov.au',
    'accept': 'text/plain, */*; q=0.01',
    'accept-language': 'en-US,en;q=0.9,id;q=0.8',
    'content-type': 'application/x-www-form-urlencoded',
    # 'cookie': 'downloadStarted=false; ASP.NET_SessionId=x1qkewrzf3bkng5fe5guccs0; __utmc=185625580; __utmz=185625580.1709167767.3.3.utmcsr=upwork.com|utmccn=(referral)|utmcmd=referral|utmcct=/; TgaSearchOption=Rto; Organisation_AjaxDetailsLoadScope_ScopeQualification=20; ifShowHistory=false; __utma=185625580.1342116690.1708677775.1709175837.1709180464.5; __utmt=1; .ASPXANONYMOUS=TZ4IATFyAr8sYnN-VyfM1-Nezh4esW6nhqc-NSyNFUuXMeeW9joXb16yxqMxN0OH1brP5IRl7edCR79tQUiYB0nu4yuqYXCcDYcZTkBqtA5dMqO_HnmfiyoluLoeDetDw5TOROO8D3NIXy54RoZllw2; __utmb=185625580.5.10.1709180464; Search_Index_gridRtoSearchResults=100',
    'origin': 'https://training.gov.au',
    'referer': 'https://training.gov.au/Search?SearchType=Rto&searchTitleOrCode=&searchTgaSubmit=Search&searchTgaSubmit=Search',
    'sec-ch-ua': '"Chromium";v="122", "Not(A:Brand";v="24", "Microsoft Edge";v="122"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
    'x-requested-with': 'XMLHttpRequest',
}

def parse(fileoutput, proxy, count):
    session = Session()
    if proxy != '':
        session.proxies.update({
            "http": proxy,
            "https": proxy
        })
        # print("proxy: ", proxy)
    if count == 0:
        strcount = '~'
    else:
        strcount = str(count)
    myTable = PrettyTable(["KEY","VALUE"])
    myTable.align ="l"
    myTable.add_row(["FILE OUTPUT: ", fileoutput])
    myTable.add_row(["PROXY: ", proxy])
    myTable.add_row(["COUNT: ", strcount])
    print(myTable)

    orgidlist = []
    for page in range(1, 100):
        params = {
            'implicitNrtScope': 'True',
            'includeUnregisteredRtosForScopeSearch': 'True',
            'includeUnregisteredRtos': 'True',
            'includeNotRtos': 'False',
            'orgSearchByNameSubmit': 'Search',
            'JavaScriptEnabled': 'true',
        }    
        data = {
            'page': page,
            'size': '1000',
            'orderBy': 'LegalPersonName-asc',
            'groupBy': '',
            'filter': '',
        }
        # breakpoint()
        
        res1 = session.post(
            'https://training.gov.au/Search/AjaxGetOrganisations',
            params=params,
            cookies=cookies,
            headers=headers,
            data=data,
        )
        alldata = json.loads(res1.text)['data']
        
        if len(alldata) == 0:
            break
        print('RTO Ids on page', page, "Saved...")
        for data in alldata:
            orgidlist.append((data['OrganisationId'], data['Codes']))

    # orgidlist = [('7cda8582-dc37-4772-90f4-a342c7e4dda6', '40898')]

    print("RTO found:", len(orgidlist), "Records")
    # breakpoint()
    wb = Workbook()
    wb.create_sheet("Sheet1")
    sheet = wb.worksheets[0]
    headersx = ("Lead Source (do not overwrite)",	"RTO Code",	"RTO training.gov.au URL",	"RTO Legal Name",	"RTO Business Name/s",	"RTO Status",	"RTO ABN",	"RTO ACN",	"RTO Type",	"RTO Website",	"RTO Regulator",	"RTO Initial Registration Date",	"RTO Current Registration Start Date",	"RTO Current Registration End Date",	"RTO Legal Authority",	"RTO Excerciser",	"RTO Scope - Qualification Code/s",	"RTO Scope - Qualification Title/s",	"RTO Scope - Unit Code/s",	"RTO Scope - Unit Title/s",	"RTO Scope - Skillset Code/s",	"RTO Scope - Skillset - Title/s",	"RTO Scope - Accredited Courses Code/s",	"RTO Scope - Accredited Courses Title/s",	"RTO Head Office Physical Address 1",	"RTO Head Office Physical Address 2",	"RTO Head Office Postal Address 1",	"RTO Head Office Postal Address 2",	"RTO Decision Type",	"RTO Decision Effective Date",	"RTO Decision End Date",	"RTO Decision Status",	"RTO Decision Review Status",	"Contact Full Name",	"Job Title",	"Contact Email",	"HubSpot Contact ID",	"Company Name",	"Contact Email + Company Name",	"RTO Website CLONE",	"Email without AT",	"RTO Website SIMPLIFIED",	"Company Name CLONE",	"Contact Email",	"Company ID",	"Contact Email | Company ID",	"Phone RAW",	"Phone LEN#",	"PHONE CONCAT (fx)",	"PHONE CONCAT LEN#",	"Phone",	"Fax RAW",	"Fax LEN#",	"Fax CONCAT (fx)",	"Fax CONCAT LEN#",	"Fax",	"Mobile RAW",	"MobileCleanup",	"Mobile LEN#",	"Mobile CONCAT (fx)",	"Mobile CONCAT LEN#",	"Mobile",	"First Name",	"Last Name",	"contactAddress RAW",	"contactAddress F/R to %",	"contactAddress L1RAW",	"contactAddress L2RAW",	"contactAddress L3RAW",	"Contact Address 1",	"postCode",	"state",	"city",	"billingZip",	"billingstate",	"billingcity",	"billingStreet",	"RTO Head Office Physical Address RAW",	"RTO Head Office Physical Address F/R to %",	"RTO Head Office Physical Address L1RAW",	"RTO Head Office Physical Address L2RAW",	"RTO Head Office Physical Address L3RAW",	"RTO Head Office Physical Address 1",	"RTO Head Office Physical Address 2",	"RTO Head Office Postal Address RAW",	"RTO Head Office Postal Address F/R to %",	"RTO Head Office Postal Address L1RAW",	"RTO Head Office Postal Address L2RAW",	"RTO Head Office Postal Address L3RAW",	"RTO Head Office Postal Address 1",	"RTO Head Office Postal Address 2")
    for idx, value in enumerate(headersx):
        sheet.cell(row=1, column=idx+1).value = value
    idx = 1
    for idxorg, data in enumerate(orgidlist):
        try:
            orgId = data[0]
            orgcode = data[1]
            print(idxorg+1, orgcode)
            res2 = session.get(f"https://training.gov.au/Organisation/Details/{orgcode}", cookies=cookies, headers=headers,)

            html = res2.content
            soup = BeautifulSoup(html, "html.parser")
            sum = soup.find("div", {"id":"rtoDetails-1"})
            fieldsets = sum.find("div", class_="fieldset").find_all("div",class_='display-row')
            code = ''
            legal_name = ''
            bname = ''
            status = ''
            abn = ''
            acn = ''
            rtype = ''
            website = ''
            for fset in fieldsets:
                try:
                    label = fset.find("div", class_="display-label").text
                except:
                    continue
                match label:
                    case 'Code:':
                        try:
                            code = fset.find("div", class_="display-field-no-width").text.strip()
                        except:
                            code = ""
                    case 'Legal name:':
                        try:
                            legal_name = fset.find("div", class_="display-field-unblocked-narrowest").text.strip()
                        except:
                            legal_name = ""
                    case 'Business name(s):':
                        try:
                            bname = fset.find("div", class_="display-field-unblocked-narrowest").text.strip()
                        except:
                            bname = ""
                    case 'Status:':
                        try:
                            status = fset.find("div", class_="display-field-no-width").text.strip()
                        except:
                            status = ""
                    case 'ABN:':
                        try:
                            abn = fset.find("div", class_="display-field-no-width").text.strip().replace("(external link)","")
                        except:
                            abn = ""
                    case 'ACN:':
                        try:
                            acn = fset.find("div", class_="display-field-unblocked-narrowest").text.strip()
                        except:
                            acn = ""
                    case 'RTO type:':
                        try:
                            rtype = fset.find("div", class_="display-field-unblocked-narrowest").text.strip()
                        except:
                            rtype = ""
                    case 'Web address:':
                        try:
                            website = fset.find("div", class_="display-field-unblocked-narrowest").text.strip().replace("(external link)","")
                        except:
                            website = ""
            if legal_name == '':
                try:
                    legal_name = data['LegalPersonNameNonCurrent']
                except:
                    legal_name = ""
            params = {
                'tabIndex': '1',
                '_': '1709176047722',
            }

            res3 = session.get(
                'https://training.gov.au/Organisation/AjaxDetailsLoadRegistration/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
            
            )    
            html = res3.content
            soup = BeautifulSoup(html, "html.parser")
            try:
                fieldsets = soup.find("div", class_="fieldset").find_all("div",class_='display-row')
            # breakpoint()
            except:
                continue
            regulator = ''
            regdate = ''
            regstart = ''
            regend = ''
            excerciser = ''
            authority = ''
            for fset in fieldsets:
                label = fset.find("div", class_="display-label-widest").text
                # print(label)
                match label:
                    case 'Registration manager:':
                        try:
                            regulator = fset.find("div", class_="display-field-no-width").text.strip().replace("(external link)","")
                        except:
                            regulator = ""
                    case 'Initial registration date:':
                        try:
                            regdate = fset.find("div", class_="display-field-no-width").text.strip()
                        except:
                            regdate = ""
                    case 'Start date:':
                        try:
                            regstart = fset.find("div", class_="display-field-no-width").text.strip()
                        except:
                            regstart = ""
                    case 'End date:':
                        try:
                            regend = fset.find("div", class_="display-field-no-width").text.strip()
                        except:
                            regend = ""
                    case 'Exerciser:':
                        try:
                            excerciser = fset.find("div", class_="display-field-unblocked-narrowest").text.strip()
                        except:
                            excerciser = ""
                    case 'Legal authority:':
                        try:
                            authority = fset.find("div", class_="display-field-no-width").text.strip()
                        except:
                            authority = ""
            # print(code,    legal_name,    bname,    status,    abn,    acn,    rtype,    website)
            # print(regulator, regdate, regstart, regend, excerciser)
            # print(legal_name)
            try:
                cdate = datetime.strptime(regdate ,'%d/%b/%Y')
            except:
                cdate = ""
            try:
                regdate = cdate.strftime('%d/%m/%Y')
            except:
                regdate = ""
            try:
                cdate = datetime.strptime(regstart ,'%d/%b/%Y')
            except:
                cdate = ""
            try:    
                regstart = cdate.strftime('%d/%m/%Y')
            except:
                regstart = ""
            try:
                cdate = datetime.strptime(regend ,'%d/%b/%Y')
            except:
                cdate = ""
            try:
                regend = cdate.strftime('%d/%m/%Y')
            except:
                regend = ""
            
            payload = {
                'page': '1',
                'size': '100',
                'orderBy': 'Code-asc',
                'groupBy': '',
                'filter': 'IsImplicit~eq~false',
            }

            response = session.post(
                'https://training.gov.au/Organisation/AjaxScopeUnit/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
                data=payload,
            )
            # breakpoint()
            unit_codestr = ''
            unit_titlestr = ''
            if response.status_code == 200:
                cleantxt = response.text.replace(':new Date(', ':"').replace('0),"','0","')
                datascope = json.loads(cleantxt)
                if len(datascope['data']) > 0:
                    unit_codestr = ", ".join([datascope['Code'] for datascope in datascope['data']])        
                    unit_titlestr =  ", ".join([datascope['Title'] for datascope in datascope['data']])

            response = session.post(
                'https://training.gov.au/Organisation/AjaxScopeQualification/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
                data=payload,
            )
            qual_codestr = ''
            qual_titlestr = ''
            if response.status_code == 200:
                cleantxt = response.text.replace(':new Date(', ':"').replace('0),"','0","')
                datascope = json.loads(cleantxt)
                if len(datascope['data']) > 0:
                    qual_codestr = ", ".join([datascope['Code'] for datascope in datascope['data']])
                    qual_titlestr =  ", ".join([datascope['Title'] for datascope in datascope['data']])

            response = session.post(
                'https://training.gov.au/Organisation/AjaxScopeSkillSet/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
                data=payload,
            )
            skill_codestr = ''
            skill_titlestr = ''
            if response.status_code == 200:
                cleantxt = response.text.replace(':new Date(', ':"').replace('0),"','0","')
                datascope = json.loads(cleantxt)
                if len(datascope['data']) > 0:
                    skill_codestr = ", ".join([datascope['Code'] for datascope in datascope['data']])
                    skill_titlestr =  ", ".join([datascope['Title'] for datascope in datascope['data']])

            response = session.post(
                'https://training.gov.au/Organisation/AjaxScopeAccreditedCourse/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
                data=payload,
            )
            course_codestr = ''
            course_titlestr = ''
            if response.status_code == 200:
                cleantxt = response.text.replace(':new Date(', ':"').replace('0),"','0","')
                datascope = json.loads(cleantxt)
                if len(datascope['data']) > 0:
                    course_codestr = ", ".join([datascope['Code'] for datascope in datascope['data']])
                    course_titlestr =  ", ".join([datascope['Title'] for datascope in datascope['data']])

            response = session.get(
                'https://training.gov.au/Organisation/AjaxDetailsLoadAddresses/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
            )    
            html = response.content
            soup = BeautifulSoup(html, "html.parser")
            # breakpoint()
            fieldsets = []
            try:
                fieldsets = soup.find("div", class_="fieldset").find_all("div",class_='display-row')
            # breakpoint()
            except:
                pass
            head_ph1 = ''
            head_ph2 = ''
            headrawph = ''
            headrawph_fr = ''
            headrawphl1r = ''
            headrawphl2r = ''
            headrawphl3r = ''

            head_pos1 = ''
            head_pos2 = ''
            headrawpos = ''
            headrawpos_fr = ''
            headrawposl1r = ''
            headrawposl2r = ''
            headrawposl3r = ''


            for fset in fieldsets:
                label = fset.find("div", class_="display-label").text
                # print(label)
                match label:
                    case 'Physical address:':
                        try:
                            head_phs =  [str(head_ph).replace("\xa0"," ").strip() for head_ph in fset.find("div", class_="display-field-no-width").contents if "<br/>" not in str(head_ph)]
                            headrawph = " ".join(head_phs)
                            headrawph_fr = "%".join(head_phs)
                            if len(head_phs) == 3:
                                head_ph1 = " ".join(head_phs[0:2])
                                headrawphl1r = head_phs[0]
                                headrawphl2r = head_phs[1]
                                headrawphl3r = head_phs[2]
                            else:
                                head_ph1 = head_phs[0]
                                headrawphl1r = head_phs[0]
                                headrawphl2r = head_phs[1]
                                headrawphl3r = ""
                            
                            head_ph2 = head_phs[-1]

                            # postalcode = str(head_phs[-1]).split(" ")[-1]
                            # state = str(head_phs[-1]).split(" ")[-2]
                            # city = " ".join(str(head_phs[-1]).split(" ")[0:-2])
                        except:
                            pass

                    case 'Postal address:':
                        try:
                            head_poss =  [str(head_pos).replace("\xa0"," ").strip() for head_pos in fset.find("div", class_="display-field-no-width").contents if "<br/>" not in str(head_pos)]
                            headrawpos = " ".join(head_poss)
                            headrawpos_fr = "%".join(head_poss)
                            if len(head_poss) == 3:
                                head_pos1 = " ".join(head_poss[0:2])
                                headrawposl1r = head_poss[0]
                                headrawposl2r = head_poss[1]
                                headrawposl3r = head_poss[2]
                            else:
                                head_pos1 = head_poss[0]
                                headrawposl1r = head_poss[0]
                                headrawposl2r = head_poss[1]
                                headrawposl3r = ""
                            
                            head_pos2 = head_poss[-1]

                            # postalcode = str(head_phs[-1]).split(" ")[-1]
                            # state = str(head_phs[-1]).split(" ")[-2]
                            # city = " ".join(str(head_phs[-1]).split(" ")[0:-2])
                        except:
                            pass

            # breakpoint()
            response = session.get(
                'https://training.gov.au/Organisation/AjaxDetailsLoadRegulatoryDecisions/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
            )    
            html = response.content
            soup = BeautifulSoup(html, "html.parser")
            trs = []
            regdec = []
            regefdate = []
            regendate = []
            regstat = []
            regrev = []

            try:
                trs = soup.find('tbody').find_all('tr')
            except:
                pass

            for tr in trs:
                # pass
                try:
                    regdec.append(tr.find_all('td')[1].text.strip())
                    regefdate.append(tr.find_all('td')[2].text.strip())
                    regendate.append(tr.find_all('td')[3].text.strip())
                    regstat.append(tr.find_all('td')[4].text.strip())
                    regrev.append(tr.find_all('td')[5].text.strip())
                except:
                    pass
                # breakpoint()

            contact_name = ""	
            contact_job = ""
            contact_email = ""
            contact_org = ""
            contact_website = ""
            contact_email_no_add = ""
            contact_website_simp = ""
            contact_org = ""
            contact_email = ""
            contact_phone = ""
            contact_phone_w_ccode = ""	
            contact_fax_w_ccode = ""
            contact_mobile = ""	
            contact_mobile_clean = ""	
            contact_mobile_w_ccode = ""	
            firstname = ""
            lastname = ""
            addressraw = ""
            addressraw_fr = ""	
            addressraw1r = ""
            addressraw2r = ""
            addressraw3r = ""
            address1 = ""
            postalcode = ""
            state = ""
            city = ""	
            
            response = session.get(
                'https://training.gov.au/Organisation/AjaxDetailsLoadContacts/{}'.format(orgId),
                params=params,
                cookies=cookies,
                headers=headers,
            )    
            html = response.content
            # breakpoint()
            soup = BeautifulSoup(html, "html.parser")

            #---------
            contactlist = []
            contacttrs = soup.find_all("div", class_="outer")
            for contacttr in contacttrs:
                trs = contacttr.find_all("div", class_="display-row")
                contact_name = ""
                contact_job = ""
                contact_email = ""
                contact_email_no_add = ""
                contact_website = ""
                contact_website_simp = ""
                contact_org = ""
                contact_phone = ""
                contact_fax = ""
                contact_mobile = ""
                contact_mobile_clean = ""
                addresses = ""
                for tr in trs:
                    label = tr.find("div", class_="display-label").text
                    contact_fax = ""
                    contact_phone_w_ccode = '=IF(AU{0}="","",IF(AV{0}=6,AU{0},IF(AV{0}=8,CONCATENATE(IFS($BW{0}="NSW","+612",$BW{0}="ACT","+612",$BW{0}="VIC","+613",$BW{0}="TAS","+613",$BW{0}="QLD","+617",$BW{0}="WA","+618",$BW{0}="SA","+618",$BW{0}="NT","+618"),AU{0}),IF(AND(AV{0}=11,LEFT(AU{0},2)="61"),CONCATENATE("+",AU{0}),IF(LEFT(AU{0},3)="+64",AU{0},IF(OR(AV{0}=9,AV{0}=10),CONCATENATE("+61",AU{0})))))))'.format(str(idx+2))
                    contact_fax_w_ccode =   '=IF(AZ{0}="","",IF(BA{0}=6,AZ{0},IF(BA{0}=8,CONCATENATE(IFS($BW{0}="NSW","+612",$BW{0}="ACT","+612",$BW{0}="VIC","+613",$BW{0}="TAS","+613",$BW{0}="QLD","+617",$BW{0}="WA","+618",$BW{0}="SA","+618",$BW{0}="NT","+618"),AZ{0}),IF(AND(BA{0}=11,LEFT(AZ{0},2)="61"),CONCATENATE("+",AZ{0}),IF(LEFT(AZ{0},3)="+64",AZ{0},IF(OR(BA{0}=9,BA{0}=10),CONCATENATE("+61",AZ{0})))))))'.format(idx+2)
                    contact_mobile_w_ccode = '=IF(BF{0}="","",IF(BG{0}=6,BF{0},IF(BG{0}=8,CONCATENATE(IFS($BW{0}="NSW","+612",$BW{0}="ACT","+612",$BW{0}="VIC","+613",$BW{0}="TAS","+613",$BW{0}="QLD","+617",$BW{0}="WA","+618",$BW{0}="SA","+618",$BW{0}="NT","+618"),BF{0}),IF(AND(BG{0}=11,LEFT(BF{0},2)="61"),CONCATENATE("+",BF{0}),IF(LEFT(BF{0},3)="+64",BF{0},IF(OR(BG{0}=9,BG{0}=10),CONCATENATE("+61",BF{0})))))))'.format(idx+2)

                    match label:
                        case 'Contact name:':
                            try:
                                contact_name = tr.find("div", class_="display-field-no-width").contents[0].strip()
                                rawname = " ".join(str(contact_name).split(" ")[1:])
                                lastname = "".join(str(rawname).split(" ")[-1])
                                firstname = " ".join(str(rawname).split(" ")[0:-1])
                            except:
                                contact_name = ""

                        case 'Job title:':
                            try:
                                contact_job = tr.find("div", class_="display-field-no-width").contents[0].strip()
                            except:
                                contact_job = ""

                        case 'Email:':
                            try:
                                contact_email = tr.find("div", class_="display-field-no-width").contents[0].strip()
                                contact_email_no_add = str(contact_email).split("@")[-1]
                                contact_website = website
                                contact_website_simp = website.replace("http://","").replace("https://","").replace("/","")
                            except:
                                contact_email = ""
                                contact_email_no_add = ""
                                contact_website = ""
                                contact_website_simp = ""
                        case 'Organisation name:':
                            try:
                                contact_org = tr.find("div", class_="display-field-no-width").contents[0].strip()
                            except:
                                contact_org = ""
                        case 'Phone:':
                            try:
                                contact_phone = tr.find("div", class_="display-field-no-width").contents[0].strip()
                                contact_phone = str(contact_phone).replace(" ","").replace("(0","").replace(")","").replace("+","")
                            except:
                                contact_phone = ""
                        case 'Fax:':
                            try:
                                contact_fax = tr.find("div", class_="display-field-no-width").contents[0].strip()
                                contact_fax = str(contact_fax).replace(" ","")
                            except:
                                contact_fax = ""
                        case 'Mobile:':
                            try:
                                contact_mobile = tr.find("div", class_="display-field-no-width").contents[0].strip()
                                contact_mobile_clean = str(contact_mobile).replace(" ","").replace("(", "").replace(")","")
                                if contact_mobile_clean[0] == "+" or contact_mobile_clean[0] == "0":
                                    contact_mobile_clean = contact_mobile_clean[1:]

                            except:
                                contact_mobile = ""
                                contact_mobile_clean = ""


                        case 'Address:':
                            
                            try:
                                addresses =  [str(head_ph).replace("\xa0"," ").strip() for head_ph in fset.find("div", class_="display-field-no-width").contents if "<br/>" not in str(head_ph)]
                                addressraw = " ".join(addresses)
                                addressraw_fr = "%".join(addresses)
                                if len(addresses) == 3:
                                    address1 = " ".join(addresses[0:2])
                                    addressraw1r = addresses[0]
                                    addressraw2r = addresses[1]
                                    addressraw3r = addresses[2]
                                else:
                                    address1 = addresses[0]
                                    addressraw1r = addresses[0]
                                    addressraw2r = addresses[1]
                                    addressraw3r = ""
                                
                                address2 = addresses[-1]

                                postalcode = str(addresses[-1]).split(" ")[-1]
                                state = str(addresses[-1]).split(" ")[-2]
                                city = " ".join(str(addresses[-1]).split(" ")[0:-2])
                            except:
                                pass
                contactlist.append({"contact_name": contact_name,
                                    "contact_email": contact_email,
                                    "contact_phone": contact_phone,
                                    "contact_fax": contact_fax,
                                    "contact_job": contact_job,
                                    "contact_mobile": contact_mobile,
                                    "contact_org": contact_org,
                                    "contact_website": contact_website,
                                    "addresses": addresses,
                                        
                    })

            # method 1
            # contactlist = list({x['contact_phone']:x for x in contactlist}.values())
            # contactlist = [x for x in contactlist if x['contact_phone'] != '']

            #method 2
            tmpcontactlist, v = [], set()
            for d in contactlist:
                k = d["contact_phone"] + d["contact_mobile"]
                if not k in v:
                    v.add(k)
                    if d['contact_phone'] != '' or d['contact_mobile'] != '':
                        tmpcontactlist.append(d)
            contactlist = tmpcontactlist.copy()


            idx -= 1
            for contact in contactlist:            
                contact_phone_w_ccode = '=IF(AU{0}="","",IF(AV{0}=6,AU{0},IF(AV{0}=8,CONCATENATE(IFS($BW{0}="NSW","+612",$BW{0}="ACT","+612",$BW{0}="VIC","+613",$BW{0}="TAS","+613",$BW{0}="QLD","+617",$BW{0}="WA","+618",$BW{0}="SA","+618",$BW{0}="NT","+618"),AU{0}),IF(AND(AV{0}=11,LEFT(AU{0},2)="61"),CONCATENATE("+",AU{0}),IF(LEFT(AU{0},3)="+64",AU{0},IF(OR(AV{0}=9,AV{0}=10),CONCATENATE("+61",AU{0})))))))'.format(str(idx+2))
                contact_fax_w_ccode =   '=IF(AZ{0}="","",IF(BA{0}=6,AZ{0},IF(BA{0}=8,CONCATENATE(IFS($BW{0}="NSW","+612",$BW{0}="ACT","+612",$BW{0}="VIC","+613",$BW{0}="TAS","+613",$BW{0}="QLD","+617",$BW{0}="WA","+618",$BW{0}="SA","+618",$BW{0}="NT","+618"),AZ{0}),IF(AND(BA{0}=11,LEFT(AZ{0},2)="61"),CONCATENATE("+",AZ{0}),IF(LEFT(AZ{0},3)="+64",AZ{0},IF(OR(BA{0}=9,BA{0}=10),CONCATENATE("+61",AZ{0})))))))'.format(idx+2)
                contact_mobile_w_ccode = '=IF(BF{0}="","",IF(BG{0}=6,BF{0},IF(BG{0}=8,CONCATENATE(IFS($BW{0}="NSW","+612",$BW{0}="ACT","+612",$BW{0}="VIC","+613",$BW{0}="TAS","+613",$BW{0}="QLD","+617",$BW{0}="WA","+618",$BW{0}="SA","+618",$BW{0}="NT","+618"),BF{0}),IF(AND(BG{0}=11,LEFT(BF{0},2)="61"),CONCATENATE("+",BF{0}),IF(LEFT(BF{0},3)="+64",BF{0},IF(OR(BG{0}=9,BG{0}=10),CONCATENATE("+61",BF{0})))))))'.format(idx+2)

                contact_job = contact['contact_job']
                contact_name = contact['contact_name']
                try:
                    rawname = " ".join(str(contact_name).split(" ")[1:])
                    lastname = "".join(str(rawname).split(" ")[-1])
                    firstname = " ".join(str(rawname).split(" ")[0:-1])
                except:
                    rawname = ""
                    lastname = ""
                    firstname = ""

                contact_email = contact['contact_email']

                try:
                    contact_email_no_add = str(contact_email).split("@")[-1]
                except:
                    contact_email_no_add = ""

                contact_website = website
                contact_org = contact['contact_org']
                try:
                    contact_website_simp = website.replace("http://","").replace("https://","").replace("/","")
                except:
                    contact_website_simp=""

                contact_phone = contact['contact_phone']
                try:
                    contact_phone = str(contact_phone).replace(" ","").replace("(0","").replace(")","").replace("+","")
                except:
                    pass

                contact_fax = contact['contact_fax']
                try:
                    contact_fax = str(contact_fax).replace(" ","")
                except:
                    pass
                contact_mobile = contact['contact_mobile']
                try:
                    contact_mobile_clean = str(contact_mobile).replace(" ","").replace("(", "").replace(")","")
                    if contact_mobile_clean[0] == "+" or contact_mobile_clean[0] == "0":
                        contact_mobile_clean = contact_mobile_clean[1:]
                except:
                    contact_mobile_clean = ""

                addresses = contact['addresses']
                try:
                    addressraw = " ".join(addresses)
                    addressraw_fr = "%".join(addresses)
                    if len(addresses) == 3:
                        address1 = " ".join(addresses[0:2])
                        addressraw1r = addresses[0]
                        addressraw2r = addresses[1]
                        addressraw3r = addresses[2]
                    else:
                        address1 = addresses[0]
                        addressraw1r = addresses[0]
                        addressraw2r = addresses[1]
                        addressraw3r = ""
                    
                    address2 = addresses[-1]

                    postalcode = str(addresses[-1]).split(" ")[-1]
                    state = str(addresses[-1]).split(" ")[-2]
                    city = " ".join(str(addresses[-1]).split(" ")[0:-2])
                except:
                    addressraw = ""
                    addressraw_fr = ""
                    address1=""
                    addressraw1r=""
                    addressraw2r=""
                    addressraw3r=""
                    postalcode=""
                    state=""
                    city=""

                sheet.cell(row=idx+2, column=1).value = "RTO"
                sheet.cell(row=idx+2, column=2).value = code
                sheet.cell(row=idx+2, column=3).value = f"https://training.gov.au/Organisation/Details/{orgcode}"
                sheet.cell(row=idx+2, column=4).value = legal_name
                sheet.cell(row=idx+2, column=5).value = bname
                sheet.cell(row=idx+2, column=6).value =  status
                sheet.cell(row=idx+2, column=7).value =  abn
                sheet.cell(row=idx+2, column=8).value = acn
                sheet.cell(row=idx+2, column=9).value =  rtype
                sheet.cell(row=idx+2, column=10).value =  website
                sheet.cell(row=idx+2, column=11).value =  regulator
                sheet.cell(row=idx+2, column=12).value =  regdate
                sheet.cell(row=idx+2, column=13).value =  regstart
                sheet.cell(row=idx+2, column=14).value =  regend
                sheet.cell(row=idx+2, column=15).value =  authority
                sheet.cell(row=idx+2, column=16).value =  excerciser
                sheet.cell(row=idx+2, column=17).value =   qual_codestr
                sheet.cell(row=idx+2, column=18).value =  qual_titlestr
                sheet.cell(row=idx+2, column=19).value =  unit_codestr
                sheet.cell(row=idx+2, column=20).value =  unit_titlestr
                sheet.cell(row=idx+2, column=21).value =  skill_codestr
                sheet.cell(row=idx+2, column=22).value = skill_titlestr
                sheet.cell(row=idx+2, column=23).value =  course_codestr
                sheet.cell(row=idx+2, column=24).value =  course_titlestr
                sheet.cell(row=idx+2, column=25).value =  head_ph1
                sheet.cell(row=idx+2, column=26).value =  head_ph2
                sheet.cell(row=idx+2, column=27).value =  head_pos1
                sheet.cell(row=idx+2, column=28).value =  head_pos2
                sheet.cell(row=idx+2, column=29).value =  ", ".join(regdec)
                sheet.cell(row=idx+2, column=30).value =  ", ".join(regefdate)
                sheet.cell(row=idx+2, column=31).value =  ", ".join(regendate)
                sheet.cell(row=idx+2, column=32).value =  ", ".join(regstat)
                sheet.cell(row=idx+2, column=33).value =  ", ".join(regrev)
                sheet.cell(row=idx+2, column=34).value =  contact_name
                sheet.cell(row=idx+2, column=35).value =  contact_job
                sheet.cell(row=idx+2, column=36).value =  contact_email
                sheet.cell(row=idx+2, column=37).value =  ""
                sheet.cell(row=idx+2, column=38).value =  contact_org
                sheet.cell(row=idx+2, column=39).value =  f"{contact_email} + {contact_org}"
                sheet.cell(row=idx+2, column=40).value =  contact_website
                sheet.cell(row=idx+2, column=41).value =  contact_email_no_add
                sheet.cell(row=idx+2, column=42).value =  contact_website_simp
                sheet.cell(row=idx+2, column=43).value =  contact_org
                sheet.cell(row=idx+2, column=44).value =  contact_email
                sheet.cell(row=idx+2, column=45).value =  ""
                sheet.cell(row=idx+2, column=46).value =  f"{contact_email} | "
                sheet.cell(row=idx+2, column=47).value =  contact_phone
                sheet.cell(row=idx+2, column=48).value =  f'=LEN(AU{idx+2})'
                sheet.cell(row=idx+2, column=49).value =  contact_phone_w_ccode
                sheet.cell(row=idx+2, column=50).value =  f'=LEN(AW{idx+2})'
                sheet.cell(row=idx+2, column=51).value =  f'=AW{idx+2}'
                sheet.cell(row=idx+2, column=52).value =  contact_fax
                sheet.cell(row=idx+2, column=53).value =  f'=LEN(AZ{idx+2})'
                sheet.cell(row=idx+2, column=54).value =  contact_fax_w_ccode
                sheet.cell(row=idx+2, column=55).value =  f'=LEN(BB{idx+2})'
                sheet.cell(row=idx+2, column=56).value =  f'=BB{idx+2}'
                sheet.cell(row=idx+2, column=57).value =  contact_mobile
                sheet.cell(row=idx+2, column=58).value =  contact_mobile_clean
                sheet.cell(row=idx+2, column=59).value =  f'=LEN(BF{idx+2})'
                sheet.cell(row=idx+2, column=60).value =  contact_mobile_w_ccode
                sheet.cell(row=idx+2, column=61).value =  f'=LEN(BH{idx+2})'
                sheet.cell(row=idx+2, column=62).value =  f'=BH{idx+2}'
                sheet.cell(row=idx+2, column=63).value =  firstname
                sheet.cell(row=idx+2, column=64).value =  lastname
                sheet.cell(row=idx+2, column=65).value =  addressraw
                sheet.cell(row=idx+2, column=66).value =  addressraw_fr
                sheet.cell(row=idx+2, column=67).value =  addressraw1r
                sheet.cell(row=idx+2, column=68).value =  addressraw2r
                sheet.cell(row=idx+2, column=69).value =  addressraw3r
                sheet.cell(row=idx+2, column=70).value =  address1
                sheet.cell(row=idx+2, column=71).value =  postalcode
                sheet.cell(row=idx+2, column=72).value =  state
                sheet.cell(row=idx+2, column=73).value =  city
                sheet.cell(row=idx+2, column=74).value =  postalcode
                sheet.cell(row=idx+2, column=75).value =  state
                sheet.cell(row=idx+2, column=76).value =  address1
                sheet.cell(row=idx+2, column=77).value =  address1
                sheet.cell(row=idx+2, column=78).value =  headrawph
                sheet.cell(row=idx+2, column=79).value =  headrawph_fr
                sheet.cell(row=idx+2, column=80).value =  headrawphl1r
                sheet.cell(row=idx+2, column=81).value =  headrawphl2r
                sheet.cell(row=idx+2, column=82).value =  headrawphl3r
                sheet.cell(row=idx+2, column=83).value = '=IF(CB{0}="","",IF(CD{0}<>"",CONCATENATE(CB{0}," ",CC{0}),CB{0}))'.format(idx+2)
                sheet.cell(row=idx+2, column=84).value = '=IF(CC{0}="","",IF(CD{0}<>"",CD{0},CC{0}))'.format(idx+2)
                sheet.cell(row=idx+2, column=85).value =  headrawpos
                sheet.cell(row=idx+2, column=86).value =  headrawpos_fr
                sheet.cell(row=idx+2, column=87).value =  headrawposl1r
                sheet.cell(row=idx+2, column=88).value =  headrawposl2r
                sheet.cell(row=idx+2, column=89).value =  headrawposl3r
                sheet.cell(row=idx+2, column=90).value = '=IF(CI{0}="","",IF(CK{0}<>"",CONCATENATE(CI{0}," ",CJ{0}),CI{0}))'.format(idx+2)
                sheet.cell(row=idx+2, column=91).value = '=IF(CJ{0}="","",IF(CK{0}<>"",CK{0},CJ{0}))'.format(idx+2)
                idx += 1   
        except ProxyError as e:
            print("Error Proxy:", str(e))
        except Exception as e:
            print("Error App:", str(e))

        idx += 1
        if count != 0:
            if idxorg >= count-1:
                break

    wb.save(s.RESFOLDER + os.path.sep + fileoutput)

def main():
    parser = argparse.ArgumentParser(description="RTO Scraper")
    parser.add_argument('-o', '--output', type=str,help="File Input")
    parser.add_argument('-p', '--proxy', type=str,help="Proxy")
    parser.add_argument('-c', '--count', type=str,help="Count")
    
    args = parser.parse_args()
    # breakpoint()
    if not args.output:
        print('use: python rtoscraper2.py -o <filename> -p <proxy> -c <count>')
        exit()
    proxy = ''
    count = 0
    if args.proxy:
        proxy = str(args.proxy)

    if args.count:
        count = int(args.count)    
    
    if args.output[-5:] != '.xlsx':
        print('File output have to XLSX file')
        exit()
    
    parse(fileoutput=args.output, proxy=proxy, count=count)
    
if __name__ == '__main__':
    main()
