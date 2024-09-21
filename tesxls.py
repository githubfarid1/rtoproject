from openpyxl import Workbook, load_workbook
    
wb = Workbook()

wb.create_sheet("Sheet1")
sheet = wb.worksheets[0]

headersx = ("Lead Source (do not overwrite)",	"RTO Code",	"RTO training.gov.au URL",	"RTO Legal Name",	"RTO Business Name/s",	"RTO Status",	"RTO ABN",	"RTO ACN",	"RTO Type",	"RTO Website",	"RTO Regulator",	"RTO Initial Registration Date",	"RTO Current Registration Start Date",	"RTO Current Registration End Date",	"RTO Legal Authority",	"RTO Excerciser",	"RTO Scope - Qualification Code/s",	"RTO Scope - Qualification Title/s",	"RTO Scope - Unit Code/s",	"RTO Scope - Unit Title/s",	"RTO Scope - Skillset Code/s",	"RTO Scope - Skillset - Title/s",	"RTO Scope - Accredited Courses Code/s",	"RTO Scope - Accredited Courses Title/s",	"RTO Head Office Physical Address 1",	"RTO Head Office Physical Address 2",	"RTO Head Office Postal Address 1",	"RTO Head Office Postal Address 2",	"RTO Decision Type",	"RTO Decision Effective Date",	"RTO Decision End Date",	"RTO Decision Status",	"RTO Decision Review Status",	"Contact Full Name",	"Job Title",	"Contact Email",	"HubSpot Contact ID",	"Company Name",	"Contact Email + Company Name",	"RTO Website CLONE",	"Email without AT",	"RTO Website SIMPLIFIED",	"Company Name CLONE",	"Contact Email",	"Company ID",	"Contact Email | Company ID",	"Phone RAW",	"Phone LEN#",	"PHONE CONCAT (fx)",	"PHONE CONCAT LEN#",	"Phone",	"Fax RAW",	"Fax LEN#",	"Fax CONCAT (fx)",	"Fax CONCAT LEN#",	"Fax",	"Mobile RAW",	"MobileCleanup",	"Mobile LEN#",	"Mobile CONCAT (fx)",	"Mobile CONCAT LEN#",	"Mobile",	"First Name",	"Last Name",	"contactAddress RAW",	"contactAddress F/R to %",	"contactAddress L1RAW",	"contactAddress L2RAW",	"contactAddress L3RAW",	"Contact Address 1",	"postCode",	"state",	"city",	"billingZip",	"billingstate",	"billingcity",	"billingStreet",	"RTO Head Office Physical Address RAW",	"RTO Head Office Physical Address F/R to %",	"RTO Head Office Physical Address L1RAW",	"RTO Head Office Physical Address L2RAW",	"RTO Head Office Physical Address L3RAW",	"RTO Head Office Physical Address 1",	"RTO Head Office Physical Address 2",	"RTO Head Office Postal Address RAW",	"RTO Head Office Postal Address F/R to %",	"RTO Head Office Postal Address L1RAW",	"RTO Head Office Postal Address L2RAW",	"RTO Head Office Postal Address L3RAW",	"RTO Head Office Postal Address 1",	"RTO Head Office Postal Address 2")

sheet.append(headersx)
sheet.append(headersx)
breakpoint()
wb.save('tes.xlsx')
